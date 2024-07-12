import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import numpy as np
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox, QFileDialog
from PyQt5.QtGui import QPalette, QColor, QFont
from PyQt5.QtCore import Qt
import sys

def create_excel():
    try:
        test_name = test_name_entry.text()
        num_exercises = int(num_exercises_entry.text())
        exercise_names = [name_entry.text() for name_entry in exercise_name_entries]
        exercise_points = [float(entry.text()) for entry in exercise_entries]
        
        students_data = []
        for i in range(int(num_students_entry.text())):
            student = [entry.text() for entry in student_entries[i]]
            students_data.append(student)
        
        file_path, _ = QFileDialog.getSaveFileName(window, "Shrani Excel datoteko", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = test_name

        headers = ["Priimek", "Ime"] + exercise_names + ["Skupne Točke", "Odstotek", "Ocena"]
        ws.append(headers)
        
        # Apply borders to the headers
        apply_borders(ws, 1, 1, 1, len(headers))

        max_points = sum(exercise_points)
        
        for student in students_data:
            ws.append(student)
        
        for row in range(2, ws.max_row + 1):
            scores = [ws.cell(row=row, column=col).value for col in range(3, 3 + num_exercises)]
            scores = [float(score) if score is not None else 0.0 for score in scores]  # Convert scores to float, handle empty cells
            total_points = sum(scores)
            percentage = (total_points / max_points) * 100
            
            grade = 1
            if percentage >= 90:
                grade = 5
            elif percentage >= 80:
                grade = 4
            elif percentage >= 65:
                grade = 3
            elif percentage >= 50:
                grade = 2
            
            ws.cell(row=row, column=num_exercises + 3, value=total_points)
            ws.cell(row=row, column=num_exercises + 4, value=percentage)
            ws.cell(row=row, column=num_exercises + 5, value=grade)

            # Apply borders to the row
            apply_borders(ws, row, 1, row, num_exercises + 5)
        
        # Calculate total and average points per exercise
        all_points = [sum([float(ws.cell(row=row, column=col).value or 0) for row in range(2, ws.max_row + 1)]) for col in range(3, 3 + num_exercises)]
        average_points = [np.mean([float(ws.cell(row=row, column=col).value or 0) for row in range(2, ws.max_row + 1)]) for col in range(3, 3 + num_exercises)]

        # Display total and average points in the GUI
        all_points_text = "Vse Možne Točke: " + str(max_points)
        average_points_text = "Povprečne Točke na Vajo: " + ", ".join([f"{name}: {value:.2f}" for name, value in zip(exercise_names, average_points)])
        additional_results_text.setText(f"{all_points_text}\n{average_points_text}")

        wb.save(file_path)
        
        overall_average_points = np.mean([float(ws.cell(row=row, column=num_exercises + 3).value or 0) for row in range(2, ws.max_row + 1)])
        
        result_text.setText(f"Povprečne Točke: {overall_average_points:.2f}")

    except ValueError as e:
        QMessageBox.critical(window, "Napaka pri vnosu", str(e))

def apply_borders(ws, start_row, start_col, end_row, end_col):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border

def add_exercise_entries():
    try:
        num_exercises = int(num_exercises_entry.text())
        for i in range(num_exercises):
            layout = QHBoxLayout()
            name_label = QLabel(f"Ime Vaje {i+1}:")
            name_label.setFont(QFont("Arial", 14, QFont.Bold))
            name_entry = QLineEdit()
            exercise_name_entries.append(name_entry)
            points_label = QLabel(f"Točke za Vajo {i+1}:")
            points_label.setFont(QFont("Arial", 14, QFont.Bold))
            points_entry = QLineEdit()
            exercise_entries.append(points_entry)

            name_entry.setFont(QFont("Arial", 12))
            points_entry.setFont(QFont("Arial", 12))

            layout.addWidget(name_label)
            layout.addWidget(name_entry)
            layout.addWidget(points_label)
            layout.addWidget(points_entry)
            exercise_layout.addLayout(layout)
    except ValueError:
        QMessageBox.critical(window, "Napaka pri vnosu", "Vnesite veljavno število vaj")

def add_student_entries():
    try:
        num_students = int(num_students_entry.text())
        for i in range(num_students):
            layout = QHBoxLayout()
            
            last_name_label = QLabel(f"Priimek Učenca {i+1}:")
            last_name_label.setFont(QFont("Arial", 12))
            last_name_entry = QLineEdit()
            last_name_entry.setFont(QFont("Arial", 12))
            first_name_label = QLabel("Ime:")
            first_name_label.setFont(QFont("Arial", 12))
            first_name_entry = QLineEdit()
            first_name_entry.setFont(QFont("Arial", 12))
            student_entries.append([last_name_entry, first_name_entry])
            
            layout.addWidget(last_name_label)
            layout.addWidget(last_name_entry)
            layout.addWidget(first_name_label)
            layout.addWidget(first_name_entry)
            
            for j in range(int(num_exercises_entry.text())):
                score_label = QLabel(f"Točke za Vajo {j+1}:")
                score_label.setFont(QFont("Arial", 12))
                score_entry = QLineEdit()
                score_entry.setFont(QFont("Arial", 12))
                student_entries[-1].append(score_entry)
                layout.addWidget(score_label)
                layout.addWidget(score_entry)
            
            student_layout.addLayout(layout)
    except ValueError:
        QMessageBox.critical(window, "Napaka pri vnosu", "Vnesite veljavno število učencev")

app = QApplication(sys.argv)

# Set dark theme
app.setStyle("Fusion")
dark_palette = QPalette()
dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
dark_palette.setColor(QPalette.WindowText, Qt.white)
dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
dark_palette.setColor(QPalette.ToolTipText, Qt.white)
dark_palette.setColor(QPalette.Text, Qt.white)
dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
dark_palette.setColor(QPalette.ButtonText, Qt.white)
dark_palette.setColor(QPalette.BrightText, Qt.red)
dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
dark_palette.setColor(QPalette.HighlightedText, Qt.black)
app.setPalette(dark_palette)

window = QWidget()
window.setWindowTitle("Kalkulator Testnih Točk")
window.showMaximized()

main_layout = QVBoxLayout()

font = QFont("Arial", 14, QFont.Bold)

test_name_label = QLabel("Ime Testa:")
test_name_label.setFont(font)
test_name_entry = QLineEdit()
test_name_entry.setFont(QFont("Arial", 14))
num_exercises_label = QLabel("Število Vaj:")
num_exercises_label.setFont(font)
num_exercises_entry = QLineEdit()
num_exercises_entry.setFont(QFont("Arial", 14))

exercise_name_entries = []
exercise_entries = []
exercise_layout = QVBoxLayout()
add_exercises_button = QPushButton("Dodaj Vaje")
add_exercises_button.setFont(font)
add_exercises_button.clicked.connect(add_exercise_entries)

num_students_label = QLabel("Število Učencev:")
num_students_label.setFont(font)
num_students_entry = QLineEdit()
num_students_entry.setFont(QFont("Arial", 14))
student_entries = []
student_layout = QVBoxLayout()
add_students_button = QPushButton("Dodaj Učence")
add_students_button.setFont(font)
add_students_button.clicked.connect(add_student_entries)

create_excel_button = QPushButton("Ustvari Excel Datoteko")
create_excel_button.setFont(font)
create_excel_button.clicked.connect(create_excel)

result_text = QLabel("")
result_text.setFont(font)

additional_results_text = QLabel("")
additional_results_text.setFont(font)

main_layout.addWidget(test_name_label)
main_layout.addWidget(test_name_entry)
main_layout.addWidget(num_exercises_label)
main_layout.addWidget(num_exercises_entry)
main_layout.addWidget(add_exercises_button)
main_layout.addLayout(exercise_layout)
main_layout.addWidget(num_students_label)
main_layout.addWidget(num_students_entry)
main_layout.addWidget(add_students_button)
main_layout.addLayout(student_layout)
main_layout.addWidget(create_excel_button)
main_layout.addWidget(result_text)
main_layout.addWidget(additional_results_text)

window.setLayout(main_layout)

sys.exit(app.exec_())
